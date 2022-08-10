from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook

def export_txl(dictonary, sheet, start_col=0, subject_col="kibbuts", subject_row="mekomot"):
    index_col = start_col + 1
    index_row = 1
    sheet.cell(row=index_row, column=index_col, value=subject_col).font = Font(bold=True)

    sheet.cell(row=index_row, column=index_col + 1, value=subject_row).font = Font(bold=True)
    index_row += 1
    for key, val in dictonary.items():
        sheet.cell(row=index_row, column=index_col, value=key).font = Font(bold=True)
        sheet.cell(row=index_row, column=index_col + 1, value=val).font = Font(bold=True)
        index_row += 1


def export_row_txl(dictonary, sheet, row_col=0):
    index_col = 2
    index_row = row_col + 1
    for key, val in dictonary.items():
        sheet.cell(row=index_row, column=index_col, value=val).font = Font(bold=True)
        index_col += 1


def printNice(dictonary):
    for key, val in dictonary.items():
        print(key, " --> ", val)


# creat a workbook obj
wb = Workbook()

# load existing spreadsheet
wb = load_workbook("ex.xlsx")

# creat an active worksheet
ws = wb.active

prefer = {}
used_mek = []
ans = {}
# col 1 is for the name of the object (in our case mekomot)
a = "B"
# print(chr(ord(a) + 1))
b_col = True
i_col = 0
i_row = 2
while b_col:
    # row 1 is for the name of the object (in our case k)
    i_row = 2
    b_row = True
    prefer[ws[chr(ord(a) + i_col)][0].value] = {}
    while b_row:
        if ws[i_row][0].value is None:
            b_row = False
        else:
            prefer[ws[chr(ord(a) + i_col)][0].value][ws[i_row][0].value] = [ws[i_row][i_col + 1].value]
            i_row += 1

    i_col += 1
    if ws[chr(ord(a) + i_col)][0].value is None:
        b_col = False

# sort by min value
for key, val in prefer.items():
    prefer[key] = dict(sorted(val.items(), key=lambda item: item[1]))

for key, dict_k in prefer.items():
    # print("\n\n", key)
    for mek, val in dict_k.items():
        if mek not in used_mek:
            ans[key] = mek
            used_mek.append(mek)
            break

printNice(ans)
export_txl(ans, ws, i_col + 5)
export_row_txl(ans, ws, i_row)
wb.save("ex1.xlsx")

