from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl import Workbook


def export_txl(dictonary, sheet, start_col=0, subject_col="kibbutzim", subject_row="mekomot"):
    index_col = start_col + 1
    index_row = 1
    sheet.cell(row=index_row, column=index_col, value=subject_col).font = Font(bold=True)

    sheet.cell(row=index_row, column=index_col + 1, value=subject_row).font = Font(bold=True)
    index_row += 1
    for key, val in dictonary.items():
        sheet.cell(row=index_row, column=index_col, value=key).font = Font(bold=True)
        sheet.cell(row=index_row, column=index_col + 1, value=val).font = Font(bold=True)
        index_row += 1


def printNice(dictonary):
    for key, val in dictonary.items():
        print(key, " --> ", val)


#####################     important    ###################################
###### you need to change it according to the number of the mekomot ####
number_of_mekomot = 12

# creat a workbook obj
wb = Workbook()
# load existing spreadsheet
wb = load_workbook("p_mek.xlsx")
# wb1 = wb["לפי העדפות גרעין"]
# creat an active worksheet
ws = wb.active
ws = wb["לפי העדפות קיבוץ"]

# dict to save the preference
prefer = {}
# list of "used" m
used_m = []
assign = {}

kibuzim_bed = {}
mek_head = {}

# get the information from the xl
indexC = 0
indexR = 0
# mek_head
for col in ws.iter_cols(max_col=1):
    for cell in col:
        indexR += 1
        if indexR != 1 and indexR <= number_of_mekomot + 1:
            mek_head[cell.value] = ws.cell(indexR, 2).value

# kibuzim_bed
indexC = 0
for row in ws.iter_rows(min_row=number_of_mekomot + 2):
    for cell in row:
        indexC += 1
        if indexC != 1 and indexC != 2:
            kibuzim_bed[ws.cell(1, indexC).value] = cell.value

indexC = 2
for col in ws.iter_cols(min_col=3):
    indexR = 0
    indexC += 1
    for cell in col:
        indexR += 1
        if indexR == 1:
            prefer[ws.cell(1, indexC).value] = {}
        if 1 < indexR < number_of_mekomot + 2:
            prefer[ws.cell(1, indexC).value][ws.cell(indexR, 1).value] = cell.value

catgoryI = 0
while catgoryI < number_of_mekomot:
    category = dict()
    catgoryI += 1
    for k_i_name, k_i_dict in prefer.items():
        for key, val in k_i_dict.items():
            if val == catgoryI:
                if k_i_name not in category:
                    category[k_i_name] = []
                    if key not in category[k_i_name]:
                        category[k_i_name].append(key)
                else:
                    category[k_i_name].append(key)
    category = dict(sorted(category.items(), key=lambda item: len(item[1])))
    for key, val in mek_head.items():
        candidate = []
        # for i in range(len([category.keys()][-1])):
        for m, ki in category.items():
            # if len(category[m]) == i+1:
            if key in category[m]:
                if key not in assign:
                    candidate.append((m, len(category[m])))
        # sort by head count
        candidate.sort(key=lambda x: x[1], reverse=True)
        # print(candidate)
        if len(candidate) > 0:
            for c in candidate:
                if mek_head[key] <= kibuzim_bed[c[0]]:
                    if c[0] not in used_m:
                        used_m.append(c[0])
                        assign[c[0]] = key
                        # way to not repeat myself
                        kibuzim_bed[key] = -1
                        break

printNice(assign)

export_txl(assign, ws, len(assign) + 4)
wb.save("kib_preference.xlsx")
