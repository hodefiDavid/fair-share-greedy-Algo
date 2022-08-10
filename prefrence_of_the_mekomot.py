from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl import Workbook


def export_txl(dictonary, sheet, start_col=0, subject_col="kibbutzim", subject_row="mekomot"):
    index_col = start_col + 1
    index_row = 1
    sheet.cell(row=index_row, column=index_col, value=subject_col).font = Font(bold=True)

    sheet.cell(row=index_row, column=index_col + 1, value=subject_row).font = Font(bold=True)
    index_row += 1
    for key, val in dictonary.items():
        sheet.cell(row=index_row, column=index_col, value=key).font = Font(bold=True)
        sheet.cell(row=index_row, column=index_col + 1, value=val).font = Font(bold=True)
        index_row += 1


def printNice(dictonary):
    for key, val in dictonary.items():
        print(key, " --> ", val)


#####################     important    ###################################
###### you need to change it according to the number of the kibbutzim ####
number_of_kibuzim = 11

# creat a workbook obj
wb = Workbook()
# load existing spreadsheet
wb = load_workbook("p_mek.xlsx")
# creat an active worksheet
ws = wb.active
ws = wb["לפי העדפות גרעין"]

# dict to save the preference
prefer = {}
# list of "used" m
used_m = []
assign = {}

kibuzim_bed = {}
mek_head = {}

# get the information from the xl
indexC = 0
indexR = 0
for col in ws.iter_cols():
    indexR = 0
    indexC += 1
    if indexC == 1:
        for cell in col:
            indexR += 1
            if indexR != 1 and indexR <= number_of_kibuzim + 1:
                kibuzim_bed[cell.value] = ws.cell(indexR, indexC + 1).value
                # print(cell.value)
    elif indexC != 2:
        prefer[ws.cell(1, indexC).value] = {}
        for cell in col:
            indexR += 1
            if indexR != 1 and indexR != number_of_kibuzim + 2:
                prefer[ws.cell(1, indexC).value][ws.cell(indexR, 1).value] = cell.value
            elif indexR == number_of_kibuzim + 2:
                mek_head[ws.cell(1, indexC).value] = cell.value

catgoryI = 0
while catgoryI < number_of_kibuzim:
    category = dict()
    catgoryI += 1
    for m_i_name, m_i_dict in prefer.items():
        for key, val in m_i_dict.items():
            if val == catgoryI:
                if m_i_name not in category:
                    category[m_i_name] = []
                    if key not in category[m_i_name]:
                        category[m_i_name].append(key)
                else:
                    category[m_i_name].append(key)
    category = dict(sorted(category.items(), key=lambda item: len(item[1])))

    for key, val in kibuzim_bed.items():
        candidate = []
        # for i in range(len([category.keys()][-1])):
        for m, ki in category.items():
            # if len(category[m]) == i+1:
            if key in category[m]:
                if key not in assign:
                    candidate.append((m, len(category[m])))
        # sort by head count
        candidate.sort(key=lambda x: x[1], reverse=True)
        # print(candidate)
        if len(candidate) > 0:
            for c in candidate:
                if kibuzim_bed[key] >= mek_head[c[0]]:
                    if c[0] not in used_m:
                        used_m.append(c[0])
                        assign[c[0]] = key
                        # way to not repeat myself
                        kibuzim_bed[key] = -1
                        break

# assign Raanana to the other places
for k, i in mek_head.items():
    if k not in assign:
        assign[k] = "Raanana"
        used_m.append(k)
printNice(assign)

export_txl(assign, ws, len(used_m) + 4)
wb.save("mek_preference.xlsx")
